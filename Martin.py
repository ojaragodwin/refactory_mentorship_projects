import openpyxl

# Load the products from an Excel sheet named "products"
products_workbook = openpyxl.load_workbook('products.xlsx')
products_sheet = products_workbook.active
products = {}

for row in products_sheet.iter_rows(min_row=2, values_only=True):
    product_name, price = row
    products[product_name] = price

# Load the accounts from an Excel sheet named "accounts"
accounts_workbook = openpyxl.load_workbook('accounts.xlsx')
accounts_sheet = accounts_workbook.active
accounts = {}

for row in accounts_sheet.iter_rows(min_row=2, values_only=True):
    email, password = row
    accounts[email] = password

# Prompt the buyer to create an account or log in
email = input('Email: ')
password = input('Password: ')

if email in accounts:
    if accounts[email] != password:
        print('Incorrect password')
        exit()
else:
    accounts[email] = password
    accounts_sheet.append([email, password])
    accounts_workbook.save('accounts.xlsx')

# Display the list of available products and prices
print('Products available:')

for product_name, price in products.items():
    print(f'{product_name}: {price}')

# Allow the buyer to choose what to buy
cart = {}

while True:
    product_name = input('Product name (type "checkout" to pay): ')

    if product_name == 'checkout':
        break

    if product_name not in products:
        print('Product not found')
        continue

    quantity = int(input('Quantity: '))
    cart[product_name] = cart.get(product_name, 0) + quantity

# Calculate the total price of the buyer's order
total_price = sum(products[product_name] * quantity for product_name, quantity in cart.items())

print(f'Total price: {total_price}')

# Allow the buyer to pay for their order
while True:
    amount_tendered = float(input('Amount tendered: '))
    if amount_tendered < total_price:
        print('Insufficient amount')
        continue
    else:
        break

balance = amount_tendered - total_price
print(f'Balance: {balance}')

# Generate a receipt for the buyer
print('Receipt:')
print('-' * 50)
print('Product\t\tQuantity\tPrice')

for product_name, quantity in cart.items():
    price = products[product_name] * quantity
    print(f'{product_name}\t\t{quantity}\t\t{price}')

print('-' * 50)
print(f'Total:\t\t\t\t{total_price}')
print(f'Amount tendered:\t\t{amount_tendered}')
print(f'Balance:\t\t\t{balance}')