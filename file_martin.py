import openpyxl

# Create a new workbook for the products
products_workbook = openpyxl.Workbook()
products_sheet = products_workbook.active

# Set the column headings
products_sheet.cell(row=1, column=1, value="Product Name")
products_sheet.cell(row=1, column=2, value="Price")

# Save the workbook to a file named "products.xlsx"
products_workbook.save("products.xlsx")

# Create a new workbook for the accounts
accounts_workbook = openpyxl.Workbook()
accounts_sheet = accounts_workbook.active

# Set the column headings
accounts_sheet.cell(row=1, column=1, value="Email")
accounts_sheet.cell(row=1, column=2, value="Password")

# Save the workbook to a file named "accounts.xlsx"
accounts_workbook.save("accounts.xlsx")